from __future__ import annotations

from datetime import date, datetime
import re
from pathlib import Path
from typing import Any
import warnings
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SPECIAL_EMPTY_VALUES = {"", "/", "无", "暂无", "未提供", "不涉及"}
ID_CARD_KEYWORDS = ["身份证", "身份证号", "证件号码", "居民身份证"]
PHONE_KEYWORDS = ["手机", "手机号", "联系电话", "联系人电话", "电话", "电话号码", "联系方式"]
AGE_KEYWORDS = ["\u5e74\u9f84", "\u5c81\u6570", "\u5e74\u7eaa"]


def get_cell_value(ws: Worksheet, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    if value is None:
        value = _merged_cell_value(ws, row, col)
    return _normalize_template_text(value)


def build_field_name(ws: Worksheet, header_rows: list[int], col: int) -> str:
    fallback = ""
    for row in reversed(header_rows):
        value = get_cell_value(ws, row, col)
        if not value:
            continue
        if fallback == "":
            fallback = value
        if _looks_like_field_name(value):
            return value
    return fallback


def can_parse_date(value: Any) -> bool:
    return parse_date_value(value) is not None


def parse_date_value(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    text = str(value).strip()
    for pattern, fmt in _date_parse_formats("ymd") + _date_parse_formats("ym"):
        if not re.fullmatch(pattern, text):
            continue
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def infer_date_granularity(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return "ymd"
    text = _normalize_template_text(value)
    if _looks_like_year_instruction(text):
        return "y"
    if _looks_like_year_month_instruction(text):
        return "ym"
    if re.fullmatch(r"\d{4}[-./]\d{1,2}", text):
        return "ym"
    if re.fullmatch(r"\d{4}年\d{1,2}月", text):
        return "ym"
    return "ymd"


def infer_date_format(value: Any, number_format: str | None = None, granularity: str | None = None) -> str:
    granularity = granularity or infer_date_granularity(value)
    if granularity == "y":
        return "yyyy"
    if granularity == "ym":
        text = _normalize_template_text(value)
        if text in {"year/month", "yyyy/mm", "年/月"}:
            return "yyyy/mm"
        if text in {"year-month", "yyyy-mm", "年-月"}:
            return "yyyy/mm" if text == "year-month" else "yyyy-mm"
        if text in {"year.month", "yyyy.mm", "年.月"}:
            return "yyyy.mm"
        if re.fullmatch(r"\d{4}-\d{1,2}", text):
            return "yyyy-mm" if len(text.split("-")[1]) == 2 else "yyyy-m"
        if re.fullmatch(r"\d{4}\.\d{1,2}", text):
            return "yyyy.mm" if len(text.split(".")[1]) == 2 else "yyyy.m"
        if re.fullmatch(r"\d{4}/\d{1,2}", text):
            return "yyyy/mm" if len(text.split("/")[1]) == 2 else "yyyy/m"
        if re.fullmatch(r"\d{4}年\d{1,2}月", text):
            return "yyyy年mm月" if _year_month_parts_are_padded(text) else "yyyy年m月"
        return "yyyy/mm"

    excel_format = _infer_date_format_from_excel_number_format(number_format)
    if excel_format:
        return excel_format
    if isinstance(value, (datetime, date)):
        return "yyyy/mm/dd"
    text = str(value).strip()
    if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", text):
        return "yyyy-mm-dd" if _date_parts_are_padded(text, "-") else "yyyy-m-d"
    if re.fullmatch(r"\d{4}\.\d{1,2}\.\d{1,2}", text):
        return "yyyy.mm.dd" if _date_parts_are_padded(text, ".") else "yyyy.m.d"
    if re.fullmatch(r"\d{4}/\d{1,2}/\d{1,2}", text):
        return "yyyy/mm/dd" if _date_parts_are_padded(text, "/") else "yyyy/m/d"
    if re.fullmatch(r"\d{8}", text):
        return "yyyy/mm/dd"
    if re.fullmatch(r"\d{4}年\d{1,2}月\d{1,2}日", text):
        return "yyyy年mm月dd日" if _chinese_date_parts_are_padded(text) else "yyyy年m月d日"
    return "yyyy/mm/dd"


def is_positive_integer(value: Any) -> bool:
    return re.fullmatch(r"[1-9]\d*", str(value).strip()) is not None


def is_number(value: Any) -> bool:
    try:
        float(str(value).strip())
        return True
    except ValueError:
        return False


def parse_number_with_unit(value: Any) -> dict[str, str] | None:
    text = str(value).strip()
    match = re.fullmatch(r"([0-9]+(?:\.[0-9]+)?)([\u4e00-\u9fa5A-Za-z㎡²/%]+)", text)
    if not match:
        return None
    return {"number": match.group(1), "unit": match.group(2)}


def infer_list_item_rule(items: list[str]) -> dict[str, Any]:
    cleaned_items = [item.strip().rstrip("。") for item in items if item.strip()]
    if not cleaned_items:
        return {"type": "string"}

    first = cleaned_items[0]
    unit_info = parse_number_with_unit(first)
    if unit_info and all((parse_number_with_unit(item) or {}).get("unit") == unit_info["unit"] for item in cleaned_items):
        return {"type": "number_with_unit", "unit": unit_info["unit"]}

    if all(is_positive_integer(item) for item in cleaned_items):
        return {"type": "positive_integer"}
    if all(is_number(item) for item in cleaned_items):
        return {"type": "number"}
    if all(can_parse_date(item) for item in cleaned_items):
        granularity = infer_date_granularity(first)
        return {"type": "date", "date_granularity": granularity, "date_format": infer_date_format(first, granularity=granularity)}
    return {"type": "string"}


def infer_rule(field_name: str, sample_value: Any, col_idx: int, sample_number_format: str | None = None) -> dict[str, Any]:
    field_name = str(field_name).strip()
    sample_text = _normalize_template_text(sample_value)
    rule: dict[str, Any] = {
        "col_idx": col_idx,
        "field_name": field_name,
        "sample_value": sample_text,
        "type": "string",
    }

    if _matches_keyword(sample_text, ID_CARD_KEYWORDS):
        rule["type"] = "id_card"
        return rule
    if _matches_keyword(sample_text, PHONE_KEYWORDS):
        rule["type"] = "phone"
        return rule
    if _looks_like_id_card_sample(sample_text):
        rule["type"] = "id_card"
        return rule
    if _looks_like_phone_sample(sample_text):
        rule["type"] = "phone"
        return rule
    if _looks_like_age_field(field_name, sample_text):
        rule["type"] = "age"
        return rule
    if sample_text in SPECIAL_EMPTY_VALUES:
        rule["type"] = "any"
        return rule

    if _is_date_rule_source(sample_value, sample_text):
        granularity = infer_date_granularity(sample_value)
        rule["type"] = "date"
        rule["date_granularity"] = granularity
        rule["date_format"] = infer_date_format(sample_value, sample_number_format, granularity)
        return rule

    if "/" in sample_text and sample_text != "/":
        parts = [part.strip() for part in sample_text.split("/") if part.strip()]
        if len(parts) >= 2:
            rule["type"] = "enum"
            rule["allowed_values"] = parts
            return rule

    if "；" in sample_text or ";" in sample_text:
        normalized = sample_text.replace(";", "；")
        rule["type"] = "list"
        rule["separator"] = "；"
        rule["item_rule"] = infer_list_item_rule(normalized.split("；"))
        return rule

    if "X" in sample_text.upper():
        rule["type"] = "string"
        return rule

    unit_info = parse_number_with_unit(sample_text)
    if unit_info:
        rule["type"] = "number_with_unit"
        rule["unit"] = unit_info["unit"]
        return rule

    if is_positive_integer(sample_text):
        rule["type"] = "positive_integer"
        return rule
    if is_number(sample_text):
        rule["type"] = "number"
        return rule

    return rule


def get_first_sheet_name(file_path: str | Path) -> str:
    wb = _load_workbook_safely(file_path)
    return wb.sheetnames[0]


def read_template_rules(
    template_path: str | Path,
    sheet_name: str,
    header_rows: list[int],
    sample_row: int,
) -> list[dict[str, Any]]:
    wb = _load_workbook_safely(template_path)
    ws = wb[sheet_name]
    raw_values = _load_raw_sheet_values(template_path, sheet_name)
    rules = []

    for col in range(1, ws.max_column + 1):
        field_name = build_field_name(ws, header_rows, col)
        sample_cell = ws.cell(row=sample_row, column=col)
        sample_raw_value = sample_cell.value
        if _needs_raw_xml_fallback(sample_cell):
            sample_raw_value = raw_values.get(sample_cell.coordinate, sample_raw_value)
        if sample_raw_value is None:
            sample_raw_value = _merged_cell_value(ws, sample_row, col)
        sample_text = _normalize_template_text(sample_raw_value)
        if not field_name and not sample_text:
            continue
        rules.append(infer_rule(field_name, sample_raw_value, col, sample_number_format=sample_cell.number_format))

    return rules


def _merged_cell_value(ws: Worksheet, row: int, col: int) -> Any:
    for cell_range in ws.merged_cells.ranges:
        if cell_range.min_row <= row <= cell_range.max_row and cell_range.min_col <= col <= cell_range.max_col:
            return ws.cell(row=cell_range.min_row, column=cell_range.min_col).value
    return None


def _normalize_template_text(value: Any) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).strip().split()).rstrip("。")
    for prefix in ["填报示例：", "填报示例:", "示例：", "示例:"]:
        if text.startswith(prefix):
            return text[len(prefix) :].strip()
    return text


def _matches_keyword(field_name: str, keywords: list[str]) -> bool:
    return any(keyword in field_name for keyword in keywords)


def _looks_like_age_field(field_name: str, sample_text: str) -> bool:
    return _matches_keyword(field_name, AGE_KEYWORDS) or sample_text in AGE_KEYWORDS


def _looks_like_phone_sample(value: str) -> bool:
    text = value.strip()
    return re.fullmatch(r"1[3-9][0-9Xx]{9}", text) is not None


def _looks_like_id_card_sample(value: str) -> bool:
    text = value.strip()
    return len(text) == 18 and re.fullmatch(r"[\dXx]{18}", text) is not None


def _looks_like_field_name(value: str) -> bool:
    if re.fullmatch(r"\d+(?:\.\d+)?", value):
        return False
    if "/" in value:
        return False
    if len(value) > 20:
        return False
    return True


def _looks_like_date_instruction(value: str) -> bool:
    return value in {"日期", "具体日期", "年月日", "年/月/日", "年-月-日", "年.月.日"}


def _looks_like_year_month_instruction(value: str) -> bool:
    return value in {"年月", "年/月", "年-月", "年.月", "year-month", "year/month", "year.month", "yyyy-mm", "yyyy/mm", "yyyy.mm"}


def _looks_like_year_instruction(value: str) -> bool:
    return value in {"year", "yyyy", "\u5e74", "\u5e74\u4efd"}


def _is_date_rule_source(sample_value: Any, sample_text: str) -> bool:
    if isinstance(sample_value, (datetime, date)):
        return True
    if _looks_like_date_instruction(sample_text) or _looks_like_year_month_instruction(sample_text) or _looks_like_year_instruction(sample_text):
        return True
    if re.fullmatch(r"\d{8}", sample_text):
        return parse_date_value(sample_text) is not None
    if re.fullmatch(r"\d{4}[-./]\d{1,2}[-./]\d{1,2}", sample_text):
        return parse_date_value(sample_text) is not None
    if re.fullmatch(r"\d{4}年\d{1,2}月\d{1,2}日", sample_text):
        return parse_date_value(sample_text) is not None
    if re.fullmatch(r"\d{4}[-./]\d{1,2}", sample_text):
        return parse_date_value(sample_text) is not None
    if re.fullmatch(r"\d{4}年\d{1,2}月", sample_text):
        return parse_date_value(sample_text) is not None
    return False


def _date_parse_formats(granularity: str) -> list[tuple[str, str]]:
    if granularity == "y":
        return [
            (r"\d{4}", "%Y"),
        ]
    if granularity == "ym":
        return [
            (r"\d{4}-\d{1,2}", "%Y-%m"),
            (r"\d{4}\.\d{1,2}", "%Y.%m"),
            (r"\d{4}/\d{1,2}", "%Y/%m"),
            (r"\d{6}", "%Y%m"),
            (r"\d{4}年\d{1,2}月", "%Y年%m月"),
        ]
    return [
        (r"\d{4}-\d{1,2}-\d{1,2}", "%Y-%m-%d"),
        (r"\d{4}\.\d{1,2}\.\d{1,2}", "%Y.%m.%d"),
        (r"\d{4}/\d{1,2}/\d{1,2}", "%Y/%m/%d"),
        (r"\d{8}", "%Y%m%d"),
        (r"\d{4}年\d{1,2}月\d{1,2}日", "%Y年%m月%d日"),
        (r"\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}:\d{1,2}", "%Y-%m-%d %H:%M:%S"),
    ]


def _date_parts_are_padded(value: str, separator: str) -> bool:
    _, month, day = value.split(separator)
    return len(month) == 2 and len(day) == 2


def _chinese_date_parts_are_padded(value: str) -> bool:
    match = re.fullmatch(r"\d{4}年(\d{1,2})月(\d{1,2})日", value)
    return bool(match and len(match.group(1)) == 2 and len(match.group(2)) == 2)


def _year_month_parts_are_padded(value: str) -> bool:
    match = re.fullmatch(r"\d{4}年(\d{1,2})月", value)
    return bool(match and len(match.group(1)) == 2)


def _load_workbook_safely(file_path: str | Path):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r"Cell .* is marked as a date but the serial value .* is outside the limits for dates.*",
            category=UserWarning,
        )
        return load_workbook(file_path, data_only=True)


def _needs_raw_xml_fallback(cell) -> bool:
    return cell.value == "#VALUE!" and cell.data_type == "e" and _looks_like_excel_date_number_format(cell.number_format)


def _looks_like_excel_date_number_format(number_format: str | None) -> bool:
    if not number_format:
        return False
    fmt = number_format.lower()
    if fmt in {"general", "@"}:
        return False
    return "y" in fmt and "m" in fmt and "d" in fmt


def _load_raw_sheet_values(workbook_path: str | Path, sheet_name: str) -> dict[str, str]:
    try:
        with ZipFile(workbook_path) as archive:
            shared_strings = _load_shared_strings(archive)
            sheet_path = _find_sheet_path(archive, sheet_name)
            if not sheet_path:
                return {}
            sheet_root = ET.fromstring(archive.read(sheet_path))
            namespace = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            values = {}
            for cell in sheet_root.findall(".//a:c", namespace):
                coordinate = cell.attrib.get("r")
                if not coordinate:
                    continue
                raw_value = _raw_cell_text(cell, shared_strings, namespace)
                if raw_value is not None:
                    values[coordinate] = raw_value
            return values
    except (KeyError, OSError, ET.ParseError):
        return {}


def _find_sheet_path(archive: ZipFile, sheet_name: str) -> str | None:
    workbook_namespace = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    relationship_namespace = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
    office_relationship_key = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root.findall("r:Relationship", relationship_namespace)}
    for sheet in workbook_root.findall("a:sheets/a:sheet", workbook_namespace):
        if sheet.attrib.get("name") != sheet_name:
            continue
        target = rel_targets.get(sheet.attrib.get(office_relationship_key))
        if not target:
            return None
        if target.startswith("/"):
            return target.lstrip("/")
        if target.startswith("xl/"):
            return target
        return f"xl/{target}"
    return None


def _load_shared_strings(archive: ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    namespace = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    strings = []
    for item in root.findall("a:si", namespace):
        texts = [text.text or "" for text in item.findall(".//a:t", namespace)]
        strings.append("".join(texts))
    return strings


def _raw_cell_text(cell: ET.Element, shared_strings: list[str], namespace: dict[str, str]) -> str | None:
    value = cell.find("a:v", namespace)
    if value is None or value.text is None:
        return None
    if cell.attrib.get("t") == "s":
        try:
            return shared_strings[int(value.text)]
        except (ValueError, IndexError):
            return value.text
    return value.text


def _infer_date_format_from_excel_number_format(number_format: str | None) -> str | None:
    if not number_format:
        return None
    fmt = number_format.lower()
    if fmt in {"general", "@"}:
        return None
    if ":" in fmt or "h" in fmt or "s" in fmt:
        return None
    if "y" not in fmt or "m" not in fmt or "d" not in fmt:
        return None
    if fmt.find("y") > fmt.find("m") or fmt.find("y") > fmt.find("d"):
        return None
    if "年" in fmt and "月" in fmt and "日" in fmt:
        return "yyyy年mm月dd日" if "mm" in fmt and "dd" in fmt else "yyyy年m月d日"
    if "/" in fmt:
        return "yyyy/mm/dd" if "mm" in fmt and "dd" in fmt else "yyyy/m/d"
    if "." in fmt:
        return "yyyy.mm.dd" if "mm" in fmt and "dd" in fmt else "yyyy.m.d"
    if "-" in fmt:
        return "yyyy-mm-dd" if "mm" in fmt and "dd" in fmt else "yyyy-m-d"
    return None
