"""
基于推断规则生成带 Excel 数据验证约束的填写模板。

输出文件保留原始模板的表头结构、示例行、样式和布局，只在数据填写区域施加：
  1. DataValidation — 输入时拦截格式错误
  2. ConditionalFormatting — 已填单元格标红（即使粘贴/绕过校验）
  3. 枚举值的下拉选项存放在隐藏 Sheet 中（规避内联长度限制）

约束尽量与 value_validator.py 的校验逻辑对齐，确保模板级拦截与 Python
级严格校验行为一致。
"""

from __future__ import annotations

from copy import copy
from pathlib import Path
import warnings
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from auto_matcher import load_auto_match_config, parse_candidates
from settings import AUTO_MATCH_CONFIG_PATH
from template_reader import SPECIAL_EMPTY_VALUES as TEMPLATE_SPECIAL_EMPTY_VALUES


ERROR_FILL = PatternFill(fill_type="solid", fgColor="FF9999")

# ── 枚举参照隐藏 Sheet 名称 ───────────────────────────────

ENUM_SHEET_NAME = "_枚举参照"

# ==============================================================
#  公开入口
# ==============================================================


def generate_fill_template(
    template_path: str | Path,
    output_path: str | Path,
    rules: list[dict[str, Any]],
    sheet_name: str,
    header_rows: list[int],
    sample_row: int,
    num_data_rows: int = 500,
    auto_match_config_path: str | Path | None = None,
) -> None:
    """
    生成带约束的 Excel 填写模板。

    Args:
        template_path: 原始模板文件路径
        output_path: 输出文件路径
        rules: 规则列表
        sheet_name: Sheet 名称
        header_rows: 表头行号 [1,2,3]
        sample_row: 示例值行号（通常 = max(header_rows) + 1）
        num_data_rows: 预留数据行数
    """
    # 用 data_only=True 避免日期序列号 warning（如 J4=20261221）
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r"Cell .* is marked as a date but the serial value .* is outside the limits for dates.*",
            category=UserWarning,
        )
        wb = load_workbook(template_path, data_only=True)
    ws = wb[sheet_name]

    # ── 行号计算 ──
    data_start = sample_row + 1
    data_end = data_start + num_data_rows - 1
    effective_rules = _apply_auto_match_dropdown_rules(rules, auto_match_config_path)

    # ── 枚举隐藏 Sheet ──
    enum_ws = _create_enum_sheet(wb, effective_rules)

    # ── 逐列施加 DataValidation + ConditionalFormatting ──
    for rule in effective_rules:
        col = int(rule["col_idx"])
        col_letter = get_column_letter(col)
        cell_range = f"{col_letter}{data_start}:{col_letter}{data_end}"

        dv = _build_validation(rule, col_letter, data_start, enum_ws)
        if dv is not None:
            dv.add(cell_range)                     # ← 关键修复：绑定范围
            ws.add_data_validation(dv)

        cf = _build_cf_rule(rule, col_letter, data_start)
        if cf is not None:
            ws.conditional_formatting.add(cell_range, cf)

    _center_template_cells(ws, effective_rules, data_end)

    # ── 保存 ──
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ==============================================================
#  内部函数
# ==============================================================


def _apply_auto_match_dropdown_rules(
    rules: list[dict[str, Any]],
    auto_match_config_path: str | Path | None,
) -> list[dict[str, Any]]:
    config = load_auto_match_config(auto_match_config_path or AUTO_MATCH_CONFIG_PATH)
    if not config.get("enabled", False):
        return [dict(rule) for rule in rules]

    auto_rules = config.get("rules", [])
    effective_rules = []
    for rule in rules:
        copied = dict(rule)
        match_rule = _find_auto_match_rule(copied.get("field_name", ""), auto_rules)
        if match_rule is not None:
            candidates = parse_candidates(match_rule.get("candidates", []))
            if candidates:
                copied["type"] = "enum"
                copied["allowed_values"] = candidates
        effective_rules.append(copied)
    return effective_rules


def _find_auto_match_rule(field_name: str, rules: list[dict[str, Any]]) -> dict[str, Any] | None:
    text = str(field_name)
    for rule in rules:
        keywords = [str(item) for item in rule.get("field_keywords", [])]
        if any(keyword and keyword in text for keyword in keywords):
            return rule
    return None


def _build_validation(
    rule: dict[str, Any],
    col_letter: str,
    row: int,
    enum_ws: Any,
) -> DataValidation | None:
    """构建 Excel DataValidation，入参 row 为首个数据行（用于自定义公式中的相对引用）。"""
    t = rule.get("type", "string")
    ref = f"{col_letter}{row}"  # 相对引用基准格

    if t == "enum":
        allowed = rule.get("allowed_values", [])
        if not allowed:
            return None
        if enum_ws is not None:
            # 引用隐藏 Sheet 的整列
            formula1 = f"={ENUM_SHEET_NAME}!${col_letter}:${col_letter}"
        else:
            formula1 = '"' + ",".join(allowed) + '"'
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
        dv.error = f"请从下拉列表中选择有效值"
        dv.errorTitle = "枚举值错误"
        dv.prompt = f"可选值：{'/'.join(allowed)}"
        dv.promptTitle = "枚举选择"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        return dv

    if t == "age":
        normalized = _age_normalized_formula(ref)
        formula1 = (
            f'OR({ref}="",AND('
            f"ISNUMBER(VALUE({normalized})),"
            f"VALUE({normalized})>=1,"
            f"VALUE({normalized})=INT(VALUE({normalized}))"
            f"))"
        )
        dv = DataValidation(type="custom", formula1=formula1, allow_blank=True)
        dv.error = "年龄必须为正整数，可填写 20 或 20岁"
        dv.errorTitle = "年龄校验"
        dv.prompt = "示例：20 或 20岁，输出会统一为 20"
        dv.promptTitle = "年龄"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        return dv

    if t == "positive_integer":
        dv = DataValidation(
            type="custom",
            formula1=f"AND(OR({ref}=\"\",ISNUMBER({ref})),{ref}>=1,{ref}=INT({ref}))",
            allow_blank=True,
        )
        dv.error = "该字段只能填写 ≥1 的整数"
        dv.errorTitle = "正整数校验"
        dv.prompt = "示例：1, 2, 100"
        dv.promptTitle = "正整数"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        return dv

    if t == "number":
        dv = DataValidation(
            type="custom",
            formula1=f"AND(OR({ref}=\"\",ISNUMBER({ref})),{ref}>=0)",
            allow_blank=True,
        )
        dv.error = "该字段只能填写 ≥0 的数字"
        dv.errorTitle = "数值校验"
        dv.prompt = "示例：0, 3.14, 100.5"
        dv.promptTitle = "数值"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        return dv

    if t == "number_with_unit":
        unit = rule.get("unit", "")
        unit_len = len(unit)
        # 自定义公式：前半可转数值 + 末尾单位匹配
        formula1 = (
            f'AND(OR({ref}="",AND('
            f"ISNUMBER(VALUE(LEFT({ref},LEN({ref})-{unit_len}))),"
            f'RIGHT({ref},{unit_len})="{unit}")))'
        )
        dv = DataValidation(type="custom", formula1=formula1, allow_blank=True)
        dv.error = f"请填写数字+单位（{unit}）的格式，如：500{unit}"
        dv.errorTitle = "带单位数值校验"
        dv.prompt = f"示例：500{unit}"
        dv.promptTitle = "数字+单位"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        return dv

    if t == "date":
        valid_formula = _date_valid_formula(ref, rule.get("date_granularity", "ymd"))
        dv = DataValidation(
            type="custom",
            formula1=f'OR({ref}="",{valid_formula})',
            allow_blank=True,
        )
        dv.error = "请填写有效日期，支持 Excel 日期或 8 位年月日（如 20261221）"
        dv.errorTitle = "日期校验"
        dv.prompt = "示例：2026/12/21 或 20261221"
        dv.promptTitle = "日期"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        return dv

    if t == "id_card":
        # 18 位，前 17 位数字，末位数字 / X / x
        formula1 = (
            f'AND(OR({ref}="",AND('
            f"LEN({ref})=18,"
            f"ISNUMBER(VALUE(LEFT({ref},17))),"
            f'OR(ISNUMBER(VALUE(RIGHT({ref},1))),RIGHT({ref},1)="X",RIGHT({ref},1)="x"))))'
        )
        dv = DataValidation(type="custom", formula1=formula1, allow_blank=True)
        dv.error = "身份证号必须为 18 位（前 17 位数字，末位数字或 X）"
        dv.errorTitle = "身份证号校验"
        dv.prompt = "前 17 位数字，末位数字或 X"
        dv.promptTitle = "身份证号"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        return dv

    if t == "phone":
        # 11 位，1 开头，第二位 3-9
        formula1 = (
            f'AND(OR({ref}="",AND('
            f"LEN({ref})=11,"
            f"ISNUMBER(VALUE(LEFT({ref},11))),"
            f'LEFT({ref},1)="1",'
            f"VALUE(MID({ref},2,1))>=3,"
            f"VALUE(MID({ref},2,1))<=9)))"
        )
        dv = DataValidation(type="custom", formula1=formula1, allow_blank=True)
        dv.error = "手机号必须为 11 位有效号码"
        dv.errorTitle = "手机号校验"
        dv.prompt = "11 位手机号，1 开头"
        dv.promptTitle = "手机号"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        return dv

    if t == "list":
        # 至少检查是否包含中文分号（多项列表）或本身无分号但已通过单项校验
        formula1 = (
            f'AND(OR({ref}="",'
            f'NOT(ISNUMBER(FIND("；",{ref}))),'
            f'ISNUMBER(FIND("；",{ref}))))'
        )
        dv = DataValidation(type="custom", formula1=formula1, allow_blank=True)
        dv.error = "多项请用中文分号（；）分隔"
        dv.errorTitle = "列表填写提示"
        dv.prompt = "示例：选项1；选项2；选项3"
        dv.promptTitle = "列表"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        return dv

    # string / any — 无约束
    return None


def _build_cf_rule(
    rule: dict[str, Any],
    col_letter: str,
    row: int,
) -> FormulaRule | None:
    """
    构建条件格式规则：单元格有值但不满足规则时 + 红色底色。

    公式逻辑：单元格非空 且 不通过规则 → 标红。
    """
    t = rule.get("type", "string")
    if t in ("string", "any"):
        return None

    ref = f"{col_letter}{row}"
    red_fill = PatternFill(fill_type="solid", fgColor="FF9999")

    if t == "enum":
        allowed = rule.get("allowed_values", [])
        if not allowed:
            return None
        # 条件格式下拉用 COUNTIF + INDIRECT 较复杂，直接用 OR(MATCH) 方式
        values_literal = ",".join(f'"{v}"' for v in allowed)
        formula = f"AND({ref}<>\"\",ISERROR(MATCH({ref},{{{values_literal}}},0)))"
        return FormulaRule(formula=[formula], fill=red_fill)

    if t == "positive_integer":
        formula = f"AND({ref}<>\"\",OR(NOT(ISNUMBER({ref})),{ref}<1,{ref}<>INT({ref})))"
        return FormulaRule(formula=[formula], fill=red_fill)

    if t == "age":
        normalized = _age_normalized_formula(ref)
        formula = (
            f'AND({ref}<>"",OR('
            f"NOT(ISNUMBER(VALUE({normalized}))),"
            f"VALUE({normalized})<1,"
            f"VALUE({normalized})<>INT(VALUE({normalized}))"
            f"))"
        )
        return FormulaRule(formula=[formula], fill=red_fill)

    if t == "number":
        formula = f"AND({ref}<>\"\",OR(NOT(ISNUMBER({ref})),{ref}<0))"
        return FormulaRule(formula=[formula], fill=red_fill)

    if t == "number_with_unit":
        unit = rule.get("unit", "")
        unit_len = len(unit)
        formula = (
            f'AND({ref}<>"",OR('
            f"NOT(ISNUMBER(VALUE(LEFT({ref},LEN({ref})-{unit_len})))),"
            f'RIGHT({ref},{unit_len})<>"{unit}"))'
        )
        return FormulaRule(formula=[formula], fill=red_fill)

    if t == "date":
        formula = f'AND({ref}<>"",NOT({_date_valid_formula(ref, rule.get("date_granularity", "ymd"))}))'
        return FormulaRule(formula=[formula], fill=red_fill)

    if t == "id_card":
        formula = (
            f'AND({ref}<>"",OR('
            f"LEN({ref})<>18,"
            f"NOT(ISNUMBER(VALUE(LEFT({ref},17)))),"
            f'AND(NOT(ISNUMBER(VALUE(RIGHT({ref},1)))),RIGHT({ref},1)<>"X",RIGHT({ref},1)<>"x")))'
        )
        return FormulaRule(formula=[formula], fill=red_fill)

    if t == "phone":
        formula = (
            f'AND({ref}<>"",OR('
            f"LEN({ref})<>11,"
            f"NOT(ISNUMBER(VALUE(LEFT({ref},11)))),"
            f'LEFT({ref},1)<>"1",'
            f"VALUE(MID({ref},2,1))<3,"
            f"VALUE(MID({ref},2,1))>9))"
        )
        return FormulaRule(formula=[formula], fill=red_fill)

    if t == "list":
        # 如果有值但不含中文分号 → 标红（要求使用分号分隔多值）
        formula = f'AND({ref}<>"",NOT(ISNUMBER(FIND("；",{ref}))))'
        return FormulaRule(formula=[formula], fill=red_fill)

    return None


def _create_enum_sheet(wb: Any, rules: list[dict[str, Any]]) -> Any:
    """创建一个隐藏 Sheet，将枚举值逐列写入，供 DataValidation 下拉引用。"""
    enum_rules = [r for r in rules if r.get("type") == "enum"]
    if not enum_rules:
        return None

    if ENUM_SHEET_NAME in wb.sheetnames:
        del wb[ENUM_SHEET_NAME]
    ws = wb.create_sheet(ENUM_SHEET_NAME)
    ws.sheet_state = "hidden"

    for rule in enum_rules:
        col = int(rule["col_idx"])
        values = rule.get("allowed_values", [])
        for i, val in enumerate(values, 1):
            ws.cell(row=i, column=col, value=val)
        # 给列宽留点余地
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width or 8,
            max((len(v) for v in values), default=0) + 2,
        )

    return ws


def _date_valid_formula(ref: str, granularity: str = "ymd") -> str:
    text = f'{ref}&""'
    year = f"VALUE(LEFT({text},4))"
    if granularity == "y":
        valid_year = (
            f"AND("
            f"LEN({text})=4,"
            f"ISNUMBER(VALUE({text})),"
            f"{year}>=1900,"
            f"{year}<=9999"
            f")"
        )
        return valid_year

    month = f"VALUE(MID({text},5,2))"
    if granularity == "ym":
        valid_compact_month = (
            f"AND("
            f"LEN({text})=6,"
            f"ISNUMBER(VALUE({text})),"
            f"{year}>=1900,"
            f"{month}>=1,"
            f"{month}<=12"
            f")"
        )
        return _allow_until_now_text(ref, valid_compact_month)

    day = f"VALUE(RIGHT({text},2))"
    valid_excel_date = f"AND(ISNUMBER({ref}),{ref}>=DATE(1900,1,1),{ref}<=DATE(9999,12,31))"
    valid_compact_date = (
        f"AND("
        f"LEN({text})=8,"
        f"ISNUMBER(VALUE({text})),"
        f"{year}>=1900,"
        f"{month}>=1,"
        f"{month}<=12,"
        f"{day}>=1,"
        f"{day}<=DAY(EOMONTH(DATE({year},{month},1),0))"
        f")"
    )
    return _allow_until_now_text(ref, f"OR({valid_excel_date},{valid_compact_date})")


def _allow_until_now_text(ref: str, formula: str) -> str:
    return f'OR({ref}="\u81f3\u4eca",{formula})'


def _age_normalized_formula(ref: str) -> str:
    text = f'{ref}&""'
    return f'IF(RIGHT({text},1)="岁",LEFT({text},LEN({text})-1),{text})'


def _center_template_cells(ws: Any, rules: list[dict[str, Any]], data_end: int) -> None:
    max_col = max(int(rule["col_idx"]) for rule in rules) if rules else ws.max_column
    max_row = max(ws.max_row, data_end)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            alignment = copy(cell.alignment)
            alignment.horizontal = "center"
            alignment.vertical = "center"
            cell.alignment = alignment
