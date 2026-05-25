from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from auto_matcher import apply_auto_match, load_auto_match_config
from settings import AUTO_MATCH_CONFIG_PATH
from value_validator import validate_by_rule


ERROR_FILL = PatternFill(fill_type="solid", fgColor="FF9999")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
NORMALIZED_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
DATA_FONT_COLOR = "FF000000"


def default_auto_match_config_path() -> Path:
    return AUTO_MATCH_CONFIG_PATH
SEQ_FIELD_NAMES = {"序号"}


def find_seq_col(rules: list[dict[str, Any]], ws=None) -> int:
    for rule in rules:
        if _looks_like_sequence_header(rule.get("field_name", "")):
            return int(rule["col_idx"])
    if ws is not None:
        header_col = find_seq_col_from_sheet(ws)
        if header_col is not None:
            return header_col
    return 1


def find_data_start_row(ws, seq_col: int) -> int:
    for row in range(1, ws.max_row + 1):
        value = get_cell_value_with_merged(ws, row, seq_col)
        if is_sequence_start_value(value):
            return row

    fallback_row = find_first_data_row_after_sequence_header(ws, seq_col)
    if fallback_row is not None:
        return fallback_row

    raise ValueError(build_data_start_not_found_message(ws, seq_col))


def get_cell_value_with_merged(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for cell_range in ws.merged_cells.ranges:
        if cell_range.min_row <= row <= cell_range.max_row and cell_range.min_col <= col <= cell_range.max_col:
            return ws.cell(row=cell_range.min_row, column=cell_range.min_col).value
    return None


def is_non_anchor_merged_cell(cell) -> bool:
    return isinstance(cell, MergedCell)


def is_sequence_start_value(value: Any) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return float(value) == 1.0

    text = str(value).strip().replace(" ", "")
    if text in {"1", "1.0", "01"}:
        return True
    if text.startswith("="):
        formula = text.upper()
        return "MAX(" in formula and formula.endswith("+1")
    return False


def find_seq_col_from_sheet(ws, max_scan_rows: int = 20):
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        for col in range(1, ws.max_column + 1):
            if _looks_like_sequence_header(get_cell_value_with_merged(ws, row, col)):
                return col
    return None


def find_first_data_row_after_sequence_header(ws, seq_col: int):
    header_row = find_sequence_header_row(ws, seq_col)
    if header_row is None:
        return None
    for row in range(header_row + 1, ws.max_row + 1):
        if not is_row_empty_with_merged(ws, row):
            return row
    return None


def find_sequence_header_row(ws, seq_col: int):
    for row in range(1, min(ws.max_row, 20) + 1):
        if _looks_like_sequence_header(get_cell_value_with_merged(ws, row, seq_col)):
            return row
    return None


def is_row_empty_with_merged(ws, row: int) -> bool:
    for col in range(1, ws.max_column + 1):
        value = get_cell_value_with_merged(ws, row, col)
        if value is not None and str(value).strip():
            return False
    return True


def _looks_like_sequence_header(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return any(name in text for name in SEQ_FIELD_NAMES)


def build_data_start_not_found_message(ws, seq_col: int) -> str:
    col_letter = get_column_letter(seq_col)
    samples = []
    for row in range(1, min(ws.max_row, 20) + 1):
        value = get_cell_value_with_merged(ws, row, seq_col)
        samples.append(f"{col_letter}{row}={value!r}")
    return (
        "没有找到序号=1的数据起始行；"
        f"当前序号列：{col_letter}；"
        f"前20行序号列值：{'; '.join(samples)}"
    )


def is_empty_row(ws, row: int, max_col: int) -> bool:
    for col in range(1, max_col + 1):
        value = get_cell_value_with_merged(ws, row, col)
        if value is not None and str(value).strip():
            return False
    return True


def normalize_data_font_color(cell) -> None:
    font = copy(cell.font)
    font.color = DATA_FONT_COLOR
    cell.font = font


def mark_error_cell(cell, field_name: str, original_value: str, message: str) -> None:
    cell.value = "/"
    cell.fill = ERROR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.comment = build_cell_comment("错误", field_name, original_value, "/", message)


def mark_warning_cell(cell, field_name: str, original_value: str, output_value: str, message: str) -> None:
    cell.fill = WARNING_FILL
    cell.comment = build_cell_comment("警告", field_name, original_value, output_value, message)


def mark_normalized_cell(cell, field_name: str, original_value: str, output_value: str, message: str) -> None:
    cell.value = output_value
    cell.fill = NORMALIZED_FILL
    cell.comment = build_cell_comment("自动修正", field_name, original_value, output_value, message)


def build_cell_comment(
    status: str,
    field_name: str,
    original_value: str,
    output_value: str,
    message: str,
) -> Comment:
    text = (
        f"{status}：{message}\n"
        f"字段：{field_name}\n"
        f"原始值：{original_value}\n"
        f"输出值：{output_value}"
    )
    return Comment(text, "simple-excel-checker")


def create_detail_sheet(wb, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(headers)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for row in rows:
        ws.append(row)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22


def validate_excel_and_write_output(
    data_path: str | Path,
    output_path: str | Path,
    rules: list[dict[str, Any]],
    sheet_name: str,
    auto_match_config_path: str | Path | None = None,
) -> dict[str, int]:
    wb = load_workbook(data_path)
    ws = wb[sheet_name]
    auto_match_config = load_auto_match_config(auto_match_config_path or default_auto_match_config_path())

    seq_col = find_seq_col(rules, ws)
    data_start_row = find_data_start_row(ws, seq_col)
    max_col = max(int(rule["col_idx"]) for rule in rules)

    error_rows = []
    warning_rows = []
    normalized_rows = []

    for row in range(data_start_row, ws.max_row + 1):
        if is_empty_row(ws, row, max_col):
            continue

        for rule in rules:
            col = int(rule["col_idx"])
            cell = ws.cell(row=row, column=col)
            if is_non_anchor_merged_cell(cell):
                continue
            normalize_data_font_color(cell)
            excel_col = get_column_letter(col)
            field_name = rule.get("field_name", "")
            cell_value = get_cell_value_with_merged(ws, row, col)
            validation = validate_by_rule(cell_value, rule)
            if validation["status"] != "ERROR":
                match_result = apply_auto_match(validation["output_value"], field_name, auto_match_config)
                if match_result["status"] in {"NORMALIZED", "WARNING", "ERROR"}:
                    if validation["status"] != "PASS":
                        match_result = dict(match_result)
                        match_result["original_value"] = validation["original_value"]
                        match_result["message"] = f"{validation['message']}；{match_result['message']}"
                    validation = match_result

            status = validation["status"]
            if status == "PASS":
                continue

            original_value = validation["original_value"]
            output_value = validation["output_value"]
            message = validation["message"]

            if status == "NORMALIZED":
                mark_normalized_cell(cell, field_name, original_value, output_value, message)
                normalized_rows.append([row, excel_col, field_name, original_value, output_value, message])
            elif status == "WARNING":
                mark_warning_cell(cell, field_name, original_value, output_value, message)
                warning_rows.append([row, excel_col, field_name, original_value, output_value, message])
            elif status == "ERROR":
                mark_error_cell(cell, field_name, original_value, message)
                error_rows.append([row, excel_col, field_name, original_value, "/", message])

    create_detail_sheet(
        wb,
        "错误明细",
        ["Excel行号", "Excel列号", "字段名", "原始值", "输出值", "错误说明"],
        error_rows,
    )
    create_detail_sheet(
        wb,
        "警告明细",
        ["Excel行号", "Excel列号", "字段名", "原始值", "输出值", "警告说明"],
        warning_rows,
    )
    create_detail_sheet(
        wb,
        "自动修正明细",
        ["Excel行号", "Excel列号", "字段名", "原始值", "修正后", "修正说明"],
        normalized_rows,
    )

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    return {
        "errors": len(error_rows),
        "warnings": len(warning_rows),
        "normalized": len(normalized_rows),
    }
