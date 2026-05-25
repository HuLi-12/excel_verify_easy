import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from output_writer import (
    default_auto_match_config_path,
    find_data_start_row,
    find_seq_col,
    get_cell_value_with_merged,
    validate_excel_and_write_output,
)
from settings import AUTO_MATCH_CONFIG_PATH


def test_find_seq_col_falls_back_to_data_sheet_header_when_rules_do_not_mark_sequence():
    wb = Workbook()
    ws = wb.active
    ws["A2"] = "\u59d3\u540d"
    ws["B2"] = "\u5e8f\u53f7"

    rules = [
        {"col_idx": 1, "field_name": "\u59d3\u540d"},
        {"col_idx": 2, "field_name": "\u5458\u5de5\u7f16\u53f7"},
    ]

    assert find_seq_col(rules, ws) == 2


def test_default_auto_match_config_path_comes_from_settings():
    assert default_auto_match_config_path() == AUTO_MATCH_CONFIG_PATH


def test_find_data_start_row_accepts_formula_sequence_start_in_merged_cell():
    wb = Workbook()
    ws = wb.active
    ws["A2"] = "\u5e8f\u53f7"
    ws["B2"] = "\u59d3\u540d"
    ws.merge_cells("A4:A12")
    ws["A4"] = "=MAX($A$1:A1)+1"
    ws["B4"] = "\u5f20\u4e09"

    assert find_data_start_row(ws, 1) == 4
    assert get_cell_value_with_merged(ws, 10, 1) == "=MAX($A$1:A1)+1"


def test_find_data_start_row_accepts_text_and_numeric_one():
    wb = Workbook()
    ws = wb.active
    ws["A2"] = "\u5e8f\u53f7"
    ws["A5"] = "01"

    assert find_data_start_row(ws, 1) == 5

    ws["A5"] = None
    ws["A6"] = 1.0

    assert find_data_start_row(ws, 1) == 6


def test_find_data_start_row_falls_back_to_first_non_empty_row_after_sequence_header():
    wb = Workbook()
    ws = wb.active
    ws["C2"] = "\u5e8f\u53f7"
    ws["D2"] = "\u59d3\u540d"
    ws["D4"] = "\u674e\u56db"

    assert find_data_start_row(ws, 3) == 4


def test_validate_excel_skips_non_anchor_merged_cells(tmp_path):
    data_path = tmp_path / "data.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A2"] = "\u5e8f\u53f7"
    ws["B2"] = "amount"
    ws.merge_cells("A4:A6")
    ws.merge_cells("B4:B6")
    ws["A4"] = "=MAX($A$1:A1)+1"
    ws["B4"] = "\u65e0"
    wb.save(data_path)

    summary = validate_excel_and_write_output(
        data_path=data_path,
        output_path=output_path,
        rules=[
            {"col_idx": 1, "field_name": "\u5e8f\u53f7", "type": "any"},
            {"col_idx": 2, "field_name": "amount", "type": "number"},
        ],
        sheet_name="Sheet1",
    )

    assert summary["warnings"] == 1


def test_validate_excel_normalizes_data_font_color_to_black(tmp_path):
    data_path = tmp_path / "data.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A2"] = "\u5e8f\u53f7"
    ws["B2"] = "name"
    ws["A4"] = 1
    ws["B4"] = "\u5f20\u4e09"
    ws["B4"].font = Font(color="FFFF0000", bold=True)
    wb.save(data_path)

    validate_excel_and_write_output(
        data_path=data_path,
        output_path=output_path,
        rules=[
            {"col_idx": 1, "field_name": "\u5e8f\u53f7", "type": "any"},
            {"col_idx": 2, "field_name": "name", "type": "string"},
        ],
        sheet_name="Sheet1",
    )

    result = load_workbook(output_path)
    cell = result["Sheet1"]["B4"]

    assert cell.font.color.rgb == "FF000000"
    assert cell.font.bold is True


def test_validate_excel_applies_enabled_auto_match_rule(tmp_path):
    data_path = tmp_path / "data.xlsx"
    output_path = tmp_path / "output.xlsx"
    config_path = tmp_path / "auto_match_rules.json"

    config_path.write_text(
        json.dumps(
            {
                "enabled": True,
                "rules": [
                    {
                        "field_keywords": ["\u516c\u53f8"],
                        "candidates": ["\u5c0f\u7c73\u516c\u53f8", "\u534e\u4e3a\u516c\u53f8"],
                        "auto_replace_threshold": 0.5,
                        "warning_threshold": 0.35,
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A2"] = "\u5e8f\u53f7"
    ws["B2"] = "\u6240\u5c5e\u516c\u53f8"
    ws["A4"] = 1
    ws["B4"] = "\u5c0f\u5c0f\u7c73"
    wb.save(data_path)

    summary = validate_excel_and_write_output(
        data_path=data_path,
        output_path=output_path,
        rules=[
            {"col_idx": 1, "field_name": "\u5e8f\u53f7", "type": "any"},
            {"col_idx": 2, "field_name": "\u6240\u5c5e\u516c\u53f8", "type": "string"},
        ],
        sheet_name="Sheet1",
        auto_match_config_path=config_path,
    )

    result = load_workbook(output_path)

    assert summary["normalized"] == 1
    assert result["Sheet1"]["B4"].value == "\u5c0f\u7c73\u516c\u53f8"


def test_validate_excel_marks_unmatched_candidate_value_as_error(tmp_path):
    data_path = tmp_path / "data.xlsx"
    output_path = tmp_path / "output.xlsx"
    config_path = tmp_path / "auto_match_rules.json"

    config_path.write_text(
        json.dumps(
            {
                "enabled": True,
                "rules": [
                    {
                        "field_keywords": ["\u516c\u53f8"],
                        "candidates": ["\u5c0f\u7c73\u516c\u53f8", "\u534e\u4e3a\u516c\u53f8"],
                        "auto_replace_threshold": 0.88,
                        "min_score_gap": 0.12,
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A2"] = "\u5e8f\u53f7"
    ws["B2"] = "\u6240\u5c5e\u516c\u53f8"
    ws["A4"] = 1
    ws["B4"] = "\u5b8c\u5168\u4e0d\u76f8\u5173"
    wb.save(data_path)

    summary = validate_excel_and_write_output(
        data_path=data_path,
        output_path=output_path,
        rules=[
            {"col_idx": 1, "field_name": "\u5e8f\u53f7", "type": "any"},
            {"col_idx": 2, "field_name": "\u6240\u5c5e\u516c\u53f8", "type": "string"},
        ],
        sheet_name="Sheet1",
        auto_match_config_path=config_path,
    )

    result = load_workbook(output_path)
    cell = result["Sheet1"]["B4"]

    assert summary["errors"] == 1
    assert cell.value == "/"
    assert cell.alignment.horizontal == "center"
    assert cell.alignment.vertical == "center"
