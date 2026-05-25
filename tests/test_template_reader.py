from datetime import datetime

from openpyxl import Workbook

from template_reader import build_field_name, infer_rule, read_template_rules
from value_validator import validate_by_rule


def test_integer_sample_stays_positive_integer_even_for_area_field():
    rule = infer_rule("建筑面积（平方米）", "23500", 1)
    assert rule["type"] == "positive_integer"


def test_field_name_slash_does_not_create_enum_without_sample_slash():
    rule = infer_rule("资产来源_划转/购入/自建", "购入", 1)
    assert rule["type"] == "string"


def test_sample_slash_always_creates_enum():
    rule = infer_rule("管理责任单位", "南昌市X投集团/XX公司", 1)
    assert rule["type"] == "enum"
    assert rule["allowed_values"] == ["南昌市X投集团", "XX公司"]


def test_phone_and_id_card_are_inferred_from_sample_text_only():
    assert infer_rule("联系电话_手机", "139XXXXXXXX", 1)["type"] == "phone"
    assert infer_rule("任意字段", "联系电话", 1)["type"] == "phone"
    assert infer_rule("任意字段", "身份证号", 1)["type"] == "id_card"


def test_phone_and_id_card_are_inferred_from_sample_shape():
    assert infer_rule("任意字段", "13800000000", 1)["type"] == "phone"
    assert infer_rule("任意字段", "11010519491231002X", 1)["type"] == "id_card"
    assert infer_rule("任意字段", "510xxxxxxxxxxxxxxx", 1)["type"] == "id_card"


def test_build_field_name_uses_plain_field_name_not_full_path():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "二、空间与物理指标"
    ws["A2"] = "有产权面积"
    ws["A3"] = "建筑面积（平方米）"

    assert build_field_name(ws, [1, 2, 3], 1) == "建筑面积（平方米）"


def test_build_field_name_ignores_option_like_header_value():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "一、资产基础法律属性"
    ws["A2"] = "资产来源"
    ws["A3"] = "划转/购入/自建"

    assert build_field_name(ws, [1, 2, 3], 1) == "资产来源"


def test_slash_date_sample_is_date_not_enum():
    rule = infer_rule("合同日期", "2025/2/3", 1)
    assert rule["type"] == "date"
    assert rule["date_format"] == "yyyy/m/d"


def test_date_instruction_sample_is_date():
    assert infer_rule("任意字段", "日期", 1)["type"] == "date"
    assert infer_rule("任意字段", "年月日", 1)["type"] == "date"


def test_datetime_sample_is_date_before_string_conversion():
    rule = infer_rule("具体地址", datetime(2026, 12, 21), 1)
    assert rule["type"] == "date"
    assert rule["date_format"] == "yyyy/mm/dd"


def test_read_template_rules_uses_raw_sample_value_for_date_inference(tmp_path):
    workbook_path = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "类别"
    ws["A2"] = "字段"
    ws["A3"] = "具体地址"
    ws["A4"] = datetime(2026, 12, 21)
    wb.save(workbook_path)

    rules = read_template_rules(workbook_path, "资产基本信息调研表", [1, 2, 3], 4)

    assert rules[0]["field_name"] == "具体地址"
    assert rules[0]["sample_value"] == "2026-12-21 00:00:00"
    assert rules[0]["type"] == "date"
    assert rules[0]["date_format"] == "yyyy/mm/dd"


def test_read_template_rules_uses_explicit_excel_date_number_format(tmp_path):
    workbook_path = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "类别"
    ws["A2"] = "字段"
    ws["A3"] = "日期字段"
    ws["A4"] = datetime(2026, 12, 21)
    ws["A4"].number_format = "yyyy/m/d"
    wb.save(workbook_path)

    rules = read_template_rules(workbook_path, "资产基本信息调研表", [1, 2, 3], 4)

    assert rules[0]["type"] == "date"
    assert rules[0]["date_format"] == "yyyy/m/d"


def test_date_instruction_uses_default_slash_date_format():
    rule = infer_rule("任意字段", "日期", 1)
    assert rule["date_format"] == "yyyy/mm/dd"


def test_compact_yyyymmdd_sample_uses_default_slash_date_format():
    rule = infer_rule("任意字段", "20261221", 1)
    assert rule["type"] == "date"
    assert rule["date_format"] == "yyyy/mm/dd"


def test_read_template_rules_falls_back_to_raw_xml_for_invalid_excel_date_serial(tmp_path):
    workbook_path = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "类别"
    ws["A2"] = "字段"
    ws["A3"] = "日期字段"
    ws["A4"] = 20261221
    ws["A4"].number_format = "mm-dd-yy"
    wb.save(workbook_path)

    rules = read_template_rules(workbook_path, "资产基本信息调研表", [1, 2, 3], 4)

    assert rules[0]["sample_value"] == "20261221"
    assert rules[0]["type"] == "date"
    assert rules[0]["date_format"] == "yyyy/mm/dd"


def test_compact_yyyymmdd_value_formats_to_template_slash_date():
    rule = infer_rule("合同日期", "2025/2/3", 1)
    validation = validate_by_rule("20250203", rule)
    assert validation["status"] == "NORMALIZED"
    assert validation["output_value"] == "2025/2/3"


def test_compact_yyyymm_value_is_cleared_for_full_date_template():
    rule = infer_rule("合同日期", "2025/2/3", 1)
    validation = validate_by_rule("202502", rule)
    assert validation["status"] == "ERROR"
    assert validation["output_value"] == "/"


def test_explicit_year_month_instruction_sets_year_month_granularity():
    rule = infer_rule("period", "year-month", 1)

    assert rule["type"] == "date"
    assert rule["date_granularity"] == "ym"
    assert rule["date_format"] == "yyyy/mm"


def test_explicit_year_month_sample_sets_year_month_granularity():
    rule = infer_rule("period", "2026/12", 1)

    assert rule["type"] == "date"
    assert rule["date_granularity"] == "ym"
    assert rule["date_format"] == "yyyy/mm"


def test_six_digit_sample_is_not_date_without_explicit_year_month_rule():
    rule = infer_rule("period", "201212", 1)

    assert rule["type"] == "positive_integer"


def test_year_month_rule_formats_compact_year_month_value():
    rule = infer_rule("period", "2026/12", 1)
    validation = validate_by_rule("201212", rule)

    assert validation["status"] == "NORMALIZED"
    assert validation["output_value"] == "2012/12"


def test_year_month_rule_rejects_full_date_value():
    rule = infer_rule("period", "2026/12", 1)
    validation = validate_by_rule("2012/12/01", rule)

    assert validation["status"] == "ERROR"
    assert validation["output_value"] == "/"


def test_year_month_and_full_date_rules_allow_until_now_text():
    until_now = "\u81f3\u4eca"

    full_date_rule = infer_rule("period", "2026/12/21", 1)
    year_month_rule = infer_rule("period", "2026/12", 1)

    assert validate_by_rule(until_now, full_date_rule)["status"] == "PASS"
    assert validate_by_rule(until_now, full_date_rule)["output_value"] == until_now
    assert validate_by_rule(until_now, year_month_rule)["status"] == "PASS"
    assert validate_by_rule(until_now, year_month_rule)["output_value"] == until_now


def test_year_only_rule_does_not_allow_until_now_text():
    rule = infer_rule("period", "year", 1)
    validation = validate_by_rule("\u81f3\u4eca", rule)

    assert validation["status"] == "ERROR"
    assert validation["output_value"] == "/"


def test_year_only_instruction_sets_year_granularity():
    rule = infer_rule("period", "year", 1)

    assert rule["type"] == "date"
    assert rule["date_granularity"] == "y"
    assert rule["date_format"] == "yyyy"


def test_year_only_rule_accepts_four_digit_year_only():
    rule = infer_rule("period", "year", 1)
    validation = validate_by_rule("2026", rule)

    assert validation["status"] == "PASS"
    assert validation["output_value"] == "2026"


def test_year_only_rule_rejects_year_month_value():
    rule = infer_rule("period", "year", 1)
    validation = validate_by_rule("2026/12", rule)

    assert validation["status"] == "ERROR"
    assert validation["output_value"] == "/"


def test_year_month_samples_with_dot_and_dash_are_detected_as_year_month():
    dot_rule = infer_rule("period", "2026.12", 1)
    dash_rule = infer_rule("period", "2026-12", 1)

    assert dot_rule["type"] == "date"
    assert dot_rule["date_granularity"] == "ym"
    assert dot_rule["date_format"] == "yyyy.mm"
    assert dash_rule["type"] == "date"
    assert dash_rule["date_granularity"] == "ym"
    assert dash_rule["date_format"] == "yyyy-mm"


def test_age_field_infers_age_rule_and_normalizes_year_suffix():
    rule = infer_rule("\u5e74\u9f84", "20", 1)
    validation = validate_by_rule("20\u5c81", rule)

    assert rule["type"] == "age"
    assert validation["status"] == "NORMALIZED"
    assert validation["output_value"] == "20"


def test_age_rule_requires_positive_integer_after_suffix_cleanup():
    rule = infer_rule("\u5e74\u9f84", "20", 1)

    assert validate_by_rule("20", rule)["status"] == "PASS"
    assert validate_by_rule("0\u5c81", rule)["status"] == "ERROR"
    assert validate_by_rule("20.5\u5c81", rule)["status"] == "ERROR"
