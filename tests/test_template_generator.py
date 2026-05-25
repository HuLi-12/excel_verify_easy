from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import json

from template_generator import ENUM_SHEET_NAME, generate_fill_template


def test_generate_fill_template_preserves_layout_and_original_styles(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "类别"
    ws["A2"] = "字段"
    ws["A3"] = "资产来源"
    ws["A4"] = "划转/购入/自建"
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="123456")
    ws["A4"].font = Font(color="FF0000", bold=True)
    wb.save(template_path)

    rules = [
        {
            "col_idx": 1,
            "field_name": "资产来源",
            "sample_value": "划转/购入/自建",
            "type": "enum",
            "allowed_values": ["划转", "购入", "自建"],
        }
    ]

    generate_fill_template(
        template_path=template_path,
        output_path=output_path,
        rules=rules,
        sheet_name="资产基本信息调研表",
        header_rows=[1, 2, 3],
        sample_row=4,
        num_data_rows=10,
    )

    result = load_workbook(output_path)
    sheet = result["资产基本信息调研表"]

    assert sheet["A1"].fill.fgColor.rgb == "00123456"
    assert sheet["A4"].font.color.rgb == "00FF0000"
    assert sheet["A4"].font.bold is True
    assert sheet["A5"].value is None
    assert sheet["A6"].value is None
    assert sheet["A4"].comment is None


def test_generate_fill_template_binds_validation_to_column_only(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "类别"
    ws["A2"] = "字段"
    ws["A3"] = "资产来源"
    ws["A4"] = "划转/购入/自建"
    ws["B3"] = "备注"
    ws["B4"] = "/"
    wb.save(template_path)

    rules = [
        {
            "col_idx": 1,
            "field_name": "资产来源",
            "sample_value": "划转/购入/自建",
            "type": "enum",
            "allowed_values": ["划转", "购入", "自建"],
        },
        {
            "col_idx": 2,
            "field_name": "备注",
            "sample_value": "/",
            "type": "any",
        },
    ]

    generate_fill_template(
        template_path=template_path,
        output_path=output_path,
        rules=rules,
        sheet_name="资产基本信息调研表",
        header_rows=[1, 2, 3],
        sample_row=4,
        num_data_rows=10,
    )

    result = load_workbook(output_path)
    sheet = result["资产基本信息调研表"]
    validations = list(sheet.data_validations.dataValidation)

    assert len(validations) == 1
    assert str(validations[0].sqref) == "A5:A14"


def test_generate_fill_template_centers_existing_and_fill_area_cells(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "类别"
    ws["A2"] = "字段"
    ws["A3"] = "序号"
    ws["A4"] = "1"
    ws["A1"].alignment = Alignment(wrap_text=True)
    wb.save(template_path)

    rules = [
        {
            "col_idx": 1,
            "field_name": "序号",
            "sample_value": "1",
            "type": "positive_integer",
        }
    ]

    generate_fill_template(
        template_path=template_path,
        output_path=output_path,
        rules=rules,
        sheet_name="资产基本信息调研表",
        header_rows=[1, 2, 3],
        sample_row=4,
        num_data_rows=10,
    )

    result = load_workbook(output_path)
    sheet = result["资产基本信息调研表"]

    assert sheet["A1"].alignment.horizontal == "center"
    assert sheet["A1"].alignment.vertical == "center"
    assert sheet["A1"].alignment.wrap_text is True
    assert sheet["A5"].alignment.horizontal == "center"
    assert sheet["A5"].alignment.vertical == "center"
    assert sheet["A14"].alignment.horizontal == "center"
    assert sheet["A14"].alignment.vertical == "center"


def test_generate_fill_template_date_validation_rejects_bad_length_and_invalid_compact_date(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "资产基本信息调研表"
    ws["A1"] = "类别"
    ws["A2"] = "字段"
    ws["A3"] = "日期"
    ws["A4"] = "2026/12/21"
    wb.save(template_path)

    rules = [
        {
            "col_idx": 1,
            "field_name": "日期",
            "sample_value": "2026/12/21",
            "type": "date",
            "date_format": "yyyy/mm/dd",
        }
    ]

    generate_fill_template(
        template_path=template_path,
        output_path=output_path,
        rules=rules,
        sheet_name="资产基本信息调研表",
        header_rows=[1, 2, 3],
        sample_row=4,
        num_data_rows=10,
    )

    result = load_workbook(output_path)
    sheet = result["资产基本信息调研表"]
    validation = list(sheet.data_validations.dataValidation)[0]
    cf_rules = list(sheet.conditional_formatting._cf_rules.values())[0]
    cf_formula = cf_rules[0].formula[0]

    assert validation.type == "custom"
    assert str(validation.sqref) == "A5:A14"
    assert "DATE(9999,12,31)" in validation.formula1
    assert 'LEN(A5&"")=8' in validation.formula1
    assert "EOMONTH" in validation.formula1
    assert "DATE(9999,12,31)" in cf_formula
    assert 'LEN(A5&"")=8' in cf_formula


def test_generate_fill_template_year_month_date_validation_uses_six_digit_rule(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DataSheet"
    ws["A1"] = "category"
    ws["A2"] = "field"
    ws["A3"] = "period"
    ws["A4"] = "2026/12"
    wb.save(template_path)

    rules = [
        {
            "col_idx": 1,
            "field_name": "period",
            "sample_value": "2026/12",
            "type": "date",
            "date_format": "yyyy/mm",
            "date_granularity": "ym",
        }
    ]

    generate_fill_template(
        template_path=template_path,
        output_path=output_path,
        rules=rules,
        sheet_name="DataSheet",
        header_rows=[1, 2, 3],
        sample_row=4,
        num_data_rows=10,
    )

    result = load_workbook(output_path)
    sheet = result["DataSheet"]
    validation = list(sheet.data_validations.dataValidation)[0]
    cf_rules = list(sheet.conditional_formatting._cf_rules.values())[0]
    cf_formula = cf_rules[0].formula[0]

    assert validation.type == "custom"
    assert 'LEN(A5&"")=6' in validation.formula1
    assert 'LEN(A5&"")=8' not in validation.formula1
    assert 'LEN(A5&"")=6' in cf_formula
    assert 'LEN(A5&"")=8' not in cf_formula


def test_generate_fill_template_year_month_date_validation_allows_until_now_text(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DataSheet"
    ws["A1"] = "category"
    ws["A2"] = "field"
    ws["A3"] = "period"
    ws["A4"] = "2026/12"
    wb.save(template_path)

    rules = [
        {
            "col_idx": 1,
            "field_name": "period",
            "sample_value": "2026/12",
            "type": "date",
            "date_format": "yyyy/mm",
            "date_granularity": "ym",
        }
    ]

    generate_fill_template(
        template_path=template_path,
        output_path=output_path,
        rules=rules,
        sheet_name="DataSheet",
        header_rows=[1, 2, 3],
        sample_row=4,
        num_data_rows=10,
    )

    result = load_workbook(output_path)
    sheet = result["DataSheet"]
    validation = list(sheet.data_validations.dataValidation)[0]
    cf_rules = list(sheet.conditional_formatting._cf_rules.values())[0]
    cf_formula = cf_rules[0].formula[0]

    assert 'A5="\u81f3\u4eca"' in validation.formula1
    assert 'A5="\u81f3\u4eca"' in cf_formula


def test_generate_fill_template_year_only_date_validation_uses_four_digit_rule(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DataSheet"
    ws["A1"] = "category"
    ws["A2"] = "field"
    ws["A3"] = "period"
    ws["A4"] = "year"
    wb.save(template_path)

    rules = [
        {
            "col_idx": 1,
            "field_name": "period",
            "sample_value": "year",
            "type": "date",
            "date_format": "yyyy",
            "date_granularity": "y",
        }
    ]

    generate_fill_template(
        template_path=template_path,
        output_path=output_path,
        rules=rules,
        sheet_name="DataSheet",
        header_rows=[1, 2, 3],
        sample_row=4,
        num_data_rows=10,
    )

    result = load_workbook(output_path)
    sheet = result["DataSheet"]
    validation = list(sheet.data_validations.dataValidation)[0]
    cf_rules = list(sheet.conditional_formatting._cf_rules.values())[0]
    cf_formula = cf_rules[0].formula[0]

    assert validation.type == "custom"
    assert 'LEN(A5&"")=4' in validation.formula1
    assert 'LEN(A5&"")=6' not in validation.formula1
    assert 'LEN(A5&"")=8' not in validation.formula1
    assert 'LEN(A5&"")=4' in cf_formula


def test_generate_fill_template_age_validation_allows_optional_year_suffix(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DataSheet"
    ws["A1"] = "category"
    ws["A2"] = "field"
    ws["A3"] = "age"
    ws["A4"] = "20"
    wb.save(template_path)

    rules = [
        {
            "col_idx": 1,
            "field_name": "\u5e74\u9f84",
            "sample_value": "20",
            "type": "age",
        }
    ]

    generate_fill_template(
        template_path=template_path,
        output_path=output_path,
        rules=rules,
        sheet_name="DataSheet",
        header_rows=[1, 2, 3],
        sample_row=4,
        num_data_rows=10,
    )

    result = load_workbook(output_path)
    sheet = result["DataSheet"]
    validation = list(sheet.data_validations.dataValidation)[0]

    assert validation.type == "custom"
    assert 'RIGHT(A5&"",1)="\u5c81"' in validation.formula1


def test_generate_fill_template_uses_auto_match_candidates_as_dropdown(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    config_path = tmp_path / "auto_match_rules.json"
    config_path.write_text(
        json.dumps(
            {
                "enabled": True,
                "rules": [
                    {
                        "field_keywords": ["\u516c\u53f8"],
                        "candidates": ["\u5c0f\u7c73\u516c\u53f8", "\u534e\u4e3a\u516c\u53f8"],
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "DataSheet"
    ws["A1"] = "category"
    ws["A2"] = "field"
    ws["A3"] = "\u6240\u5c5e\u516c\u53f8"
    ws["A4"] = "/"
    wb.save(template_path)

    rules = [
        {
            "col_idx": 1,
            "field_name": "\u6240\u5c5e\u516c\u53f8",
            "sample_value": "/",
            "type": "string",
        }
    ]

    generate_fill_template(
        template_path=template_path,
        output_path=output_path,
        rules=rules,
        sheet_name="DataSheet",
        header_rows=[1, 2, 3],
        sample_row=4,
        num_data_rows=10,
        auto_match_config_path=config_path,
    )

    result = load_workbook(output_path)
    sheet = result["DataSheet"]
    enum_sheet = result[ENUM_SHEET_NAME]
    validation = list(sheet.data_validations.dataValidation)[0]

    assert validation.type == "list"
    assert str(validation.sqref) == "A5:A14"
    assert enum_sheet["A1"].value == "\u5c0f\u7c73\u516c\u53f8"
    assert enum_sheet["A2"].value == "\u534e\u4e3a\u516c\u53f8"
